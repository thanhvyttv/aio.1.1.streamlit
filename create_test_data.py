import random

import openpyxl
from openpyxl.utils import get_column_letter


def create_student_score_excel(
    num_students=150, filename="tong_hop_diem_hoc_sinh.xlsx"
):
    """
    Tạo một file Excel (.xlsx) tổng hợp điểm số học sinh.

    Args:
        num_students (int): Số lượng học sinh cần tạo. Mặc định là 150.
        filename (str): Tên file Excel sẽ được tạo. Mặc định là "tong_hop_diem_hoc_sinh.xlsx".
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Diem Hoc Sinh"

    # Đặt tiêu đề cho các cột
    headers = ["STT", "Họ Tên", "Điểm số"]
    sheet.append(headers)

    # Tạo dữ liệu cho học sinh
    for i in range(1, num_students + 1):
        # Tạo tên học sinh đơn giản (có thể tùy chỉnh phức tạp hơn nếu cần)
        ho_ten = f"Học sinh {i:03d}"
        # Tạo điểm số ngẫu nhiên từ 0 đến 100
        diem_so = random.randint(0, 100)
        sheet.append([i, ho_ten, diem_so])

    # Tự động điều chỉnh độ rộng cột
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # Lưu file Excel
    workbook.save(filename)
    print(f"File '{filename}' đã được tạo thành công với {num_students} học sinh.")


if __name__ == "__main__":
    create_student_score_excel()
